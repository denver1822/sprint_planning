from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.models import TaskItem, VotingRound
from app.services.rooms import _require_participant, get_room_or_404


async def export_tasks_xlsx(session: AsyncSession, code: str, token: str | None) -> bytes:
    room = await get_room_or_404(session, code)
    await _require_participant(session, room.id, token)
    tasks = (
        await session.scalars(
            select(TaskItem)
            .where(TaskItem.room_id == room.id)
            .options(selectinload(TaskItem.rounds).selectinload(VotingRound.result))
            .order_by(TaskItem.position)
        )
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Оценки"
    sheet.append(["Комната", room.name])
    sheet.append(["Статус", room.state])
    sheet.append([])
    headers = ["№", "Задача", "Jira", "Итоговая оценка", "Голоса"]
    sheet.append(headers)
    for task in tasks:
        revealed_rounds = [round_ for round_ in task.rounds if round_.result is not None]
        last_revealed_round = max(revealed_rounds, key=lambda round_: round_.sequence, default=None)
        revealed_votes = " · ".join(vote["card_value"] for vote in last_revealed_round.result.revealed_votes) if last_revealed_round else ""
        sheet.append(
            [
                task.position + 1,
                task.title,
                task.jira_key or "",
                task.final_estimate or "",
                revealed_votes,
            ]
        )
    header_fill = PatternFill("solid", fgColor="1C4B45")
    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A5"
    for column, width in {"A": 7, "B": 42, "C": 16, "D": 20, "E": 42}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=5, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
